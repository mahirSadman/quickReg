import gspread
from openpyxl import workbook, load_workbook
import datetime
import requests

NsuID_col = "C"
event_col = 5
list_of_lists = list()


# connect to google sheet
try:
    sa = gspread.service_account(filename="quickreg-354618-15c4ca8155fe.json")
    sh = sa.open("QuickReg")
    worksheet = sh.worksheet("Sheet1")
except:
    print("Could not connect to google sheet")

# connect to all members excel
try:
    wb = load_workbook("info/MembersList.xlsx")
    ws = wb.active
except:
    print("Could not connect to all members excel file")

# connect to unfinished list excel
try:
    unfinished_wb = load_workbook("info/UnfinishedList.xlsx")
    unfinished_ws = unfinished_wb.active
except:
    print("Could not connect to unfinished members excel file")


def load_sheet_info_to_excel():
    try:
        global list_of_lists
        list_of_lists = worksheet.get_all_values()
    except:
        print("Could not get values from google sheet")
        # print('Rows:', worksheet.row_count)
        # list_of_dicts = worksheet.get_all_records()
        # print(list_of_lists[0])
        # return worksheet.get_all_values()
    try:
        ws.delete_rows(1, ws.max_row)
        [ws.append(x) for x in list_of_lists]
        wb.save('info/MembersList.xlsx')
    except:
        print("Could not load values to excel")

    print("Successfully loaded sheet values to excel")


def registration():
    student_id = input("\n\nEnter your NSU student ID: ")
    if student_id == 's':
        run()
    else:
        try:
            all_ids = ws[NsuID_col]
        except:
            print("could not load ids from excel")

        flag = 0
        for cell in all_ids:
            if cell.value == student_id:
                row = cell.row
                flag = 1

        if flag:
            # print(row)
            # print(time)
            ct = datetime.datetime.now()
            time = "{:d}:{:02d}".format(ct.hour, ct.minute)
            try:
                ws.cell(row=row, column=event_col).value = time
                wb.save('info/MembersList.xlsx')
            except:
                print("Could not update time to excel")

            try:
                worksheet.update_cell(row, event_col, time)
                # print("Record updated")
                name = ws.cell(row=row, column=1).value
                print("Welcome " + name)
            except:
                print("Could not update time to google sheet")
                unfinished_ws.append([student_id, time])
                unfinished_wb.save('info/UnfinishedList.xlsx')
                print("Updated to unfinished list")

        else:
            print("Ooopss... your id is not in the database")
        registration()


def check_internet():
    url = "http://www.google.com"
    timeout = 5
    try:
        request = requests.get(url, timeout=timeout)
        print("Connected to the Internet")
        # return 1
    except (requests.ConnectionError, requests.Timeout) as exception:
        # return 0
        print("No internet connection.")


def run():
    go = 1
    while go:
        print("1. Load values from google sheet to excel\n2. Start registration\n3. Exit")
        task = input("Enter task no.: ")

        if task == '1':
            load_sheet_info_to_excel()
        elif task == '2':
            registration()
        elif task == '3':
            go = 0
            break
        else:
            print("Enter valid value")

run()
